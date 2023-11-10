import logging

import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import InputMediaPhoto, ReplyKeyboardRemove

from Keyboards.default import asosiy_menu

from Keyboards.inline import language, who, important

# Your Telegram API token
TOKEN = '6966701444:AAGVJAiMxYhNr18o_IETZLLtZK2aTQ9dZ_0'
# Initialize the bot and dispatcher
bot = Bot(token=TOKEN,parse_mode='HTML')
dp = Dispatcher(bot, storage=MemoryStorage())
logging.basicConfig(level=logging.INFO)

# Middleware for logging
dp.middleware.setup(LoggingMiddleware())
DATABASE_DICT = {}


class Mars(StatesGroup):
    modme = State()


# Echo handler
@dp.message_handler(commands=['start', 'help'])
async def send_welcome(message: types.Message):
    await message.answer(
        "Mars botiga xush kelibsiz! Iltimos, Til tanlang,\n\nДобро пожаловать в Mars Bot! Пожалуйста, "
        "выберите язык\n\nWelcome to Mars Bot! Please select a language",
        reply_markup=language)

    user = message.from_user.id
    my_user = 6498877955
    if user == my_user:
        pass
    else:
        await bot.send_message('6498877955', f' New user : {user}', reply_markup=ReplyKeyboardRemove())



@dp.callback_query_handler(text='Uzbek')
async def uzb_l(call: types.CallbackQuery):
    await call.message.delete()
    await call.message.answer('Iltimos, kimligingizni ko`rsating))', reply_markup=who)


@dp.callback_query_handler(text='stud')
async def student_login(call: types.CallbackQuery):
    await call.message.answer('Modme id ni kiriting: ')
    await call.message.delete()
    await Mars.modme.set()


p = 0


@dp.message_handler(content_types=types.ContentType.TEXT, state=Mars.modme)
async def texter(message: types.Message, state: FSMContext):
    id_student = message.text
    wb = openpyxl.load_workbook('students.xlsx', 'rb')
    sheet = wb['Sheet']
    users = []
    for i in range(2, sheet.max_row + 1):
        users.append(sheet.cell(row=i, column=2).value)

    if int(id_student) in users:

        bolakaylar = []
        for k in range(2, sheet.max_row + 1):
            _ = []
            for p in range(1, 5):

                _.append(sheet.cell(row=k, column=p).value)
                if _ not in bolakaylar:
                    bolakaylar.append(_)

        async def adder():
            for t in bolakaylar:
                if str(t[1]) == str(id_student):
                    t.remove(0)
                    t.append(1)
        count = 0
        for sim in bolakaylar:
            count += 1
            print(count)
            if sim[1] == int(id_student):
                if bolakaylar[count - 1][-1] == 1:
                    await message.answer('Bunday Akkaunt Oldin ro`yxatdan o`tgan')

                elif bolakaylar[count - 1][-1] == 0:
                    await message.answer('Muvaffaqiyatli kirdingiz', reply_markup=asosiy_menu)
                    DATABASE_DICT[message.from_user.id] = int(id_student)

                    await state.finish()

                    print(DATABASE_DICT)

                    await adder()

                else:
                    await message.answer('Tizimda qandaydir nosozlik!')


        await state.finish()

    else:
        await message.answer('Bunday foydalanuvchi topilmadi!')


@dp.message_handler(text='🏫О школе')
async def profil(message:types.message,state:FSMContext):
    photo = open('defaullt/img.png', 'rb')
    await message.answer_photo(photo=photo)
    await message.answer('''
    👨🏻‍💻 Hamma kasb yaxshi, xavfsizi undan yaxshi!

Havola orqali o’ting va videoni tomosha qiling:
https://www.instagram.com/p/CxYJKchiR9x/


🥊 Boks sport turini sevuvchi har bir odam, hayotida bir marotaba bo'lsa ham professional bokschi bo'lishni orzu qilgan, lekin bu sport har birimizga ham to’g'ri kelmaydi. «Mars IT School» sizning farzandingiz hayotiga kamroq xavf tug’diradigan va zamonaviy kasblarni qulay hamda qiziqarli holda taqdim etib, kelajakda o’z kasbining egasi bo'lishiga yordam beradi.

🤝 Tanlovda adashmang, farzandingiz uchun hozirdan mustahkam poydevor quring!

<a href="https://t.me/Mars_yunusobod">📩 Yunusobod filiali</a>
<a href="https://t.me/Mars_tinchlik">📩 Tinchlik filiali</a>
<a href="https://t.me/Mars_Qutbiniso">📩 Chilonzor-Qutbiniso filialil;</a>
<a href="https://t.me/Mars_chilonzor18">📩 Chilonzor 18 filiali</a>

<b>👇🏻 Hoziroq izohlarda "+" belgisini qoldiring va biz siz bilan bog'lanamiz!</b>

<b>«Mars IT School» — bu kelajak!</b>

<b>📞 78-777-77-57</b>
    ''')



@dp.message_handler(text='💥Space shop')
async def photo(message: types.Message):
    photos = [
        InputMediaPhoto(open('images/1photo.jpg', 'rb'),),
        InputMediaPhoto(open('images/2photo.jpg', 'rb')),
        InputMediaPhoto(open('images/3photo.jpg', 'rb')),
        InputMediaPhoto(open('images/4photo.jpg', 'rb')),
        InputMediaPhoto(open('images/5photo.jpg', 'rb')),
        InputMediaPhoto(open('images/6photo.jpg', 'rb')),
        InputMediaPhoto(open('images/7photo.jpg', 'rb')),
        InputMediaPhoto(open('images/8photo.jpg', 'rb')),
        InputMediaPhoto(open('images/9photo.jpg', 'rb')),
        InputMediaPhoto(open('images/10photo.jpg', 'rb')),
    ]
    await message.answer_media_group(media=photos)
    await message.answer('Выберите приз',reply_markup=important)







if __name__ == '__main__':
    from aiogram import executor
    executor.start_polling(dp, skip_updates=True)
